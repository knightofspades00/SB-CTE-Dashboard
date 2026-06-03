"""
database/import_data.py — Imports CTE data from CTE_Connections.xlsx into SQLite.
Processes three sheets:
  Sheet 1 — school-to-pathway membership (X marks in a matrix).
  Sheet 2 — sector, department, and description enrichment for each pathway.
  Sheet 3 — career connections, linking pathways to potential job titles.
NAME_MAP corrects spelling inconsistencies between sheets so records join correctly.
Run after init_db.py:  python database/import_data.py [--file path/to/file.xlsx]
"""

import argparse
import os
import sqlite3
import sys

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from dotenv import load_dotenv
load_dotenv(os.path.join(PROJECT_ROOT, ".env"))

from config.settings import Config

try:
    import openpyxl
except ImportError:
    print("openpyxl is not installed. Run: pip install openpyxl")
    sys.exit(1)

DEFAULT_XLSX = os.path.join(PROJECT_ROOT, "database", "CTE_Connections.xlsx")

SHEET1_NAME = "School Filter Step 1"
SHEET2_NAME = "Occupational Sector Skills 2"
SHEET3_NAME = "Occupational Sector Connect  3"

S2_SECTOR      = "Occupation Sectors"
S2_DEPARTMENT  = "Departments that use or are focused on these areas:"
S2_CERT        = "Relative CTE Certifications that may link into this field \u2026"
S2_DESCRIPTION = "CTE Program Description (Provided by the Dept. of Education)"
S2_SKILLS      = "Skills & Compentenices Gained"

S3_PROGRAM  = "CTE Program"
S3_CAREERS  = "Potential SB County Career Pathways"

NAME_MAP = {
    "Agriculture and Natural Resources":           "Agricultural and Natural Resources",
    "Agriscience":                                 "Agricultural Science",
    "Architectual Design":                         "Architectural Design",
    "Building and Construction":                   "Building and Construction Trades",
    "Design, Visual, and Media Arts":              "Design Visual Arts and Media",
    "Entreprenuership and Self-Employment":         "Entrepreneurship and Self Employment",
    "Farm and Argriculture":                       "Farm and Agriculture",
    "Film Study":                                  "Film Studies",
    "Heating, Ventilation, and Air Conditioning":  "Heating, Ventilation, Air Conditioning",
    "Hospitality, Tourism, and Recreation":        "Hospitality, Tourism and Recreation",
    "Manufacturing ":                              "Manufacturing",
    "Marketing, Sales, and Services":              "Marketing, Sales and Services",
    "Operations":                                  "Operations",
    "Ornamental Horiculture":                      "Ornamental Horticulture",
    "Product and Innovation and Design":           "Product Innovation and Design",
    "Structural Repair and Refinishing ":          "Structural Repair and Refinishing",
    "Systems, Diagostics, Service, and Repair":    "Systems, Diagnostics, Service and Repair",
    "Unmanned Aircraft Systems":                   "Unmanned Aircraft Systems (Drones)",
}

def canonical(name):
    """
    Normalise a pathway name by stripping whitespace and applying NAME_MAP.
    NAME_MAP corrects the many spelling/punctuation variants that appear across the three
    spreadsheet sheets so that records from different sheets join on the same string.
    """
    if name is None:
        return None
    cleaned = name.strip()
    return NAME_MAP.get(cleaned, cleaned)

def get_connection(db_path):
    """Open and return a SQLite connection with foreign-key enforcement and Row factory enabled."""
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row
    return conn

def get_or_create(conn, table, name_col, name_val, extra=None):
    """
    Return the id of an existing row matching name_val, or insert a new one and return its id.
    extra is an optional dict of additional column-value pairs to include on insert.
    """
    row = conn.execute(f"SELECT id FROM {table} WHERE {name_col} = ?", (name_val,)).fetchone()
    if row:
        return row["id"]
    cols = [name_col]
    vals = [name_val]
    if extra:
        cols += list(extra.keys())
        vals += list(extra.values())
    placeholders = ", ".join(["?"] * len(vals))
    col_list     = ", ".join(cols)
    conn.execute(f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})", vals)
    return conn.execute(f"SELECT id FROM {table} WHERE {name_col} = ?", (name_val,)).fetchone()["id"]

def import_sheet1(conn, ws, warnings):
    """
    Import the school-to-pathway matrix from Sheet 1.
    Column headers (from col 3 onward) are pathway names; an "X" in a cell means that school
    offers that pathway.  Names are normalised through canonical() before being written.
    """
    print("\n-- Sheet 1: Schools & Pathways")
    headers      = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    pathway_cols = headers[2:]
    pathway_ids  = {}
    for raw_name in pathway_cols:
        if not raw_name:
            continue
        name = canonical(raw_name)
        pid  = get_or_create(conn, "pathways", "name", name)
        pathway_ids[raw_name] = pid
    schools_added = 0
    links_added   = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        district = str(row[0]).strip() if row[0] else None
        school   = str(row[1]).strip() if row[1] else None
        if not district or not school:
            continue
        sid = get_or_create(conn, "schools", "name", school, extra={"district": district})
        schools_added += 1
        for col_idx, raw_name in enumerate(pathway_cols):
            if not raw_name:
                continue
            cell_val = row[col_idx + 2]
            if str(cell_val).strip().upper() == "X":
                pid = pathway_ids.get(raw_name)
                if pid:
                    conn.execute("INSERT OR IGNORE INTO school_pathways (school_id, pathway_id) VALUES (?, ?)", (sid, pid))
                    links_added += 1
    print(f"  {schools_added} schools, {links_added} links")

def import_sheet2(conn, ws, warnings):
    """
    Enrich existing pathway rows with sector, department, description, and skills from Sheet 2.
    Matches rows to pathways by the CTE certification name column after canonical() normalisation.
    Inserts a new pathway row if no matching one exists yet.
    """
    print("\n-- Sheet 2: Sectors & Descriptions")
    header_row = next(ws.iter_rows(max_row=1, values_only=True))
    col        = {v: i for i, v in enumerate(header_row) if v}
    enriched   = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        cert = row[col.get(S2_CERT, 2)]
        if not cert:
            continue
        name = canonical(cert)
        existing = conn.execute("SELECT id FROM pathways WHERE name = ?", (name,)).fetchone()
        if existing:
            conn.execute("UPDATE pathways SET sector=?, department=?, description=?, skills=? WHERE id=?", (
                row[col.get(S2_SECTOR, 0)],
                row[col.get(S2_DEPARTMENT, 1)],
                row[col.get(S2_DESCRIPTION, 3)],
                row[col.get(S2_SKILLS, 4)],
                existing["id"]
            ))
            enriched += 1
        else:
            conn.execute("INSERT OR IGNORE INTO pathways (name, sector, department, description, skills) VALUES (?,?,?,?,?)", (
                name,
                row[col.get(S2_SECTOR, 0)],
                row[col.get(S2_DEPARTMENT, 1)],
                row[col.get(S2_DESCRIPTION, 3)],
                row[col.get(S2_SKILLS, 4)],
            ))
    print(f"  {enriched} pathways enriched")

def import_sheet3(conn, ws, warnings):
    """
    Link pathways to career titles from Sheet 3.
    Career titles are semicolon-separated in the source cell; each becomes a row in the
    careers table and a row in pathway_careers.
    """
    print("\n-- Sheet 3: Career Connections")
    header_row = next(ws.iter_rows(max_row=1, values_only=True))
    col        = {v: i for i, v in enumerate(header_row) if v}
    programs_linked = 0
    careers_created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        program_raw = row[col[S3_PROGRAM]]
        careers_raw = row[col[S3_CAREERS]]
        if not program_raw:
            continue
        name    = canonical(program_raw)
        pathway = conn.execute("SELECT id FROM pathways WHERE name = ?", (name,)).fetchone()
        if not pathway:
            conn.execute("INSERT OR IGNORE INTO pathways (name) VALUES (?)", (name,))
            pathway = conn.execute("SELECT id FROM pathways WHERE name = ?", (name,)).fetchone()
        pid = pathway["id"]
        if careers_raw:
            for career_title in str(careers_raw).split(";"):
                career_title = career_title.strip()
                if not career_title:
                    continue
                cid = get_or_create(conn, "careers", "name", career_title)
                conn.execute("INSERT OR IGNORE INTO pathway_careers (pathway_id, career_id) VALUES (?, ?)", (pid, cid))
                careers_created += 1
        programs_linked += 1
    print(f"  {programs_linked} programs, {careers_created} career links")

def run_import(xlsx_path):
    """Load the workbook, run all three sheet importers inside a single transaction, and commit."""
    db_path = os.path.join(PROJECT_ROOT, Config.DATABASE_PATH)
    if not os.path.exists(db_path):
        print("Database not found. Run: python database/init_db.py")
        sys.exit(1)
    if not os.path.exists(xlsx_path):
        print(f"Spreadsheet not found at {xlsx_path}")
        sys.exit(1)
    print(f"Importing from: {xlsx_path}")
    wb       = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    conn     = get_connection(db_path)
    warnings = []
    try:
        import_sheet1(conn, wb[SHEET1_NAME], warnings)
        import_sheet2(conn, wb[SHEET2_NAME], warnings)
        import_sheet3(conn, wb[SHEET3_NAME], warnings)
        conn.commit()
        print("\n✓ Import complete.")
    except Exception as e:
        conn.rollback()
        print(f"\n✗ Import failed: {e}")
        raise
    finally:
        conn.close()

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", default=DEFAULT_XLSX)
    args = parser.parse_args()
    run_import(args.file)
